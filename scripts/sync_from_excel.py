"""
Sync between index JSON files and the Excel sheet.

Anime  (Sheet1): syncs Downloads + Sub/Dub -> anime_metadata.json, versions -> anime_index.json
Manga  (Sheet2): syncs versions -> index.json
Novel  (Sheet3): syncs versions -> novel_index.json

Versions flow one way: extension .js -> index JSON -> Excel.

The .js file is authoritative because its `mangayomiSources` header is the
version Mangayomi actually loads. A number typed into the sheet is a note, not
a release: if it disagrees, the sheet is corrected from the source. Letting the
sheet win instead published JustAnime as 0.2.7 in anime_index.json while
justanime.js still said 0.2.6, so the app advertised an update that did not
exist. Downloads and Sub/Dub still come *from* the sheet — that data has no
other home.

The pre-commit hook runs this automatically on every commit.
"""

import json
import re
import subprocess
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(r"C:\Users\malik\OneDrive\Documents\Mangayomi\anime extensions versions and notes.xlsx")

# Maps Excel display names -> canonical index names (anime only)
NAME_MAP = {
    "animetsu": "Animetsu",
    "myron": "MyroniX",
    "anikototv": "AniKoto",
    "hianime": "HiAnime",
    "anidap": "Anidap",
    "animeheaven": "AnimeHeaven",
    "animeparadise": "AnimeParadise",
    "justanime": "JustAnime",
    "anicove": "AniCove",
    "animekai": "AnimeKai",
}

scripts_dir = Path(__file__).resolve().parent
main_dir    = scripts_dir.parent


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
    print(f"Saved: {path.name}")


def canonical_name(raw, name_map=None):
    key = raw.strip().lower()
    if name_map and key in name_map:
        return name_map[key]
    return raw.strip()


def staged_text(rel_path):
    """File content as it will exist in the commit, i.e. the git index.

    Read from the index rather than the working tree so a half-finished version
    bump sitting unstaged does not get published. For an unmodified file the
    index matches HEAD, so this is the working tree minus uncommitted edits.
    Returns None outside a repo or when the path is untracked.
    """
    try:
        out = subprocess.run(["git", "show", f":{rel_path}"],
                             cwd=main_dir, capture_output=True, timeout=15)
    except Exception:
        return None
    return out.stdout.decode("utf-8", "replace") if out.returncode == 0 else None


def source_version(entry):
    """Version declared in the extension's own .js — what Mangayomi loads.

    Derived from sourceCodeUrl, which ends in the file's repo-relative path.
    Returns None when the path cannot be resolved or the file is not tracked,
    so an entry we cannot verify is left alone rather than blanked.
    """
    m = re.search(r"/(javascript/.+?\.js)$", entry.get("sourceCodeUrl") or "")
    if not m:
        return None
    text = staged_text(m.group(1))
    if text is None:
        return None
    # Match the header block's version, which is the first one in the file.
    m = re.search(r'"version"\s*:\s*"([0-9][0-9.]*)"', text)
    return m.group(1) if m else None


def sync_sheet(ws, index, col_name, col_version,
               col_dl=None, col_subdub=None,
               metadata=None, name_map=None,
               label=""):
    """Sync one Excel sheet against one index list. Returns (excel_changed, reports)."""

    index_by_name = {e["name"]: e for e in index}
    excel_changed = False
    reports = {"added": [], "repo_newer": [], "unknown": [], "meta": [],
               "source_wins": []}

    # Read existing Excel rows
    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row_idx, col_name).value
        if not raw:
            continue

        name    = canonical_name(str(raw), name_map)
        xl_ver  = str(ws.cell(row_idx, col_version).value).strip() \
                  if ws.cell(row_idx, col_version).value else None

        # Sync anime metadata (Downloads / Sub/Dub)
        if metadata is not None and col_dl and col_subdub:
            dl_val  = ws.cell(row_idx, col_dl).value
            sub_val = ws.cell(row_idx, col_subdub).value
            downloads = str(dl_val).strip() if dl_val else "?"
            subdub    = str(sub_val).strip() if sub_val else "?"
            if downloads.lower() in ("yes", "true"):
                downloads = "Yes"
            elif downloads.lower() in ("no", "false"):
                downloads = "No"
            old = metadata.get(name, {})
            metadata[name] = {"downloads": downloads, "subDub": subdub}
            if old != metadata[name]:
                reports["meta"].append(f"  {name}: downloads={downloads}, subDub={subdub}")

        if name not in index_by_name:
            reports["unknown"].append(f"  {raw!r} — not in {label}, skipped")
            continue

        # The .js header outranks both the index and the sheet — see module docstring.
        src_ver = source_version(index_by_name[name])
        if src_ver and src_ver != index_by_name[name]["version"]:
            reports["source_wins"].append(
                f"  {name}: index {index_by_name[name]['version']} -> {src_ver} (from .js)")
            index_by_name[name]["version"] = src_ver

        repo_ver = index_by_name[name]["version"]

        if xl_ver and xl_ver != repo_ver:
            # One-way: a differing number in the sheet is a stale edit, not a
            # release, so correct the sheet rather than republishing the index.
            ws.cell(row_idx, col_version).value = repo_ver
            ws.cell(row_idx, col_name).value = name
            reports["repo_newer"].append(f"  {name}: Excel {xl_ver} -> {repo_ver}")
            excel_changed = True
        elif xl_ver is None:
            ws.cell(row_idx, col_version).value = repo_ver
            ws.cell(row_idx, col_name).value = name
            excel_changed = True

    # Add missing extensions
    excel_names = set()
    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row_idx, col_name).value
        if raw:
            excel_names.add(canonical_name(str(raw), name_map))

    for entry in index:
        name = entry["name"]
        if name not in excel_names:
            next_row = ws.max_row + 1
            ws.cell(next_row, col_name).value = name
            ws.cell(next_row, col_version).value = entry["version"]
            if col_dl and col_subdub and metadata:
                meta = metadata.get(name, {})
                ws.cell(next_row, col_dl).value = meta.get("downloads", "?")
                ws.cell(next_row, col_subdub).value = meta.get("subDub", "?")
            else:
                # For manga/novel, write language in col_dl slot (Language column)
                ws.cell(next_row, col_dl).value = entry.get("lang", "en")
            reports["added"].append(f"  {name} (v{entry['version']})")
            excel_changed = True

    return excel_changed, reports


def sync():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Ensure all three sheets exist
    if "Manga" not in wb.sheetnames:
        ws = wb.create_sheet("Manga")
        ws.append(["Extension name", "Language", "Version number", "Notes"])
    if "Novel" not in wb.sheetnames:
        ws = wb.create_sheet("Novel")
        ws.append(["Extension name", "Language", "Version number", "Notes"])

    print(f"\nReading: {EXCEL_PATH.name}\n")

    metadata     = load_json(scripts_dir / "anime_metadata.json")
    anime_index  = load_json(main_dir / "anime_index.json")
    manga_index  = load_json(main_dir / "index.json")
    novel_index  = load_json(main_dir / "novel_index.json")

    excel_changed = False

    # --- Anime (Sheet1) ---
    ws_anime = wb["Sheet1"]
    h = [str(c.value).strip().lower() if c.value else "" for c in ws_anime[1]]
    try:
        changed, rep = sync_sheet(
            ws_anime, anime_index,
            col_name    = h.index("extension name") + 1,
            col_version = h.index("version number") + 1,
            col_dl      = h.index("downloads") + 1,
            col_subdub  = h.index("language supported") + 1,
            metadata    = metadata,
            name_map    = NAME_MAP,
            label       = "anime_index.json"
        )
        excel_changed = excel_changed or changed
        print_reports("Anime", rep)
    except ValueError as e:
        print(f"Anime sheet error: {e}")

    # --- Manga ---
    ws_manga = wb["Manga"]
    h = [str(c.value).strip().lower() if c.value else "" for c in ws_manga[1]]
    try:
        changed, rep = sync_sheet(
            ws_manga, manga_index,
            col_name    = h.index("extension name") + 1,
            col_version = h.index("version number") + 1,
            col_dl      = h.index("language") + 1,   # Language column
            label       = "index.json"
        )
        excel_changed = excel_changed or changed
        print_reports("Manga", rep)
    except ValueError as e:
        print(f"Manga sheet error: {e}")

    # --- Novel ---
    ws_novel = wb["Novel"]
    h = [str(c.value).strip().lower() if c.value else "" for c in ws_novel[1]]
    try:
        changed, rep = sync_sheet(
            ws_novel, novel_index,
            col_name    = h.index("extension name") + 1,
            col_version = h.index("version number") + 1,
            col_dl      = h.index("language") + 1,
            label       = "novel_index.json"
        )
        excel_changed = excel_changed or changed
        print_reports("Novel", rep)
    except ValueError as e:
        print(f"Novel sheet error: {e}")

    # Save
    save_json(scripts_dir / "anime_metadata.json", metadata)
    save_json(main_dir / "anime_index.json", anime_index)

    if excel_changed:
        wb.save(EXCEL_PATH)
        print(f"Saved: {EXCEL_PATH.name}")


def print_reports(label, rep):
    print(f"=== {label} ===")
    if rep["added"]:
        print("  New rows added:")
        print("\n".join(rep["added"]))
    if rep["meta"]:
        print("  Metadata updated:")
        print("\n".join(rep["meta"]))
    if rep["source_wins"]:
        print("  Index corrected from extension source:")
        print("\n".join(rep["source_wins"]))
    if rep["repo_newer"]:
        print("  Excel updated from repo:")
        print("\n".join(rep["repo_newer"]))
    if rep["unknown"]:
        print("  Skipped (removed from repo):")
        print("\n".join(rep["unknown"]))
    if not any(rep.values()):
        print("  No changes")


if __name__ == "__main__":
    sync()
